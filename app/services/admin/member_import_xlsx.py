"""会员批量导入：Excel 模板生成（含填写说明工作表）。"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.admin.member_import_parser import TEMPLATE_HEADERS

# 示例行：供租户参考填写格式
_SAMPLE_ROW: tuple[Any, ...] = (
    "张三",
    "13800138000",
    "周卡",
    "河南省新乡市红旗区某某小区1号楼",
    4,
    6,
    1,
    "2026-07-01",
    "否",
    "否",
    "示例备注，导入前请删除本行",
)

_INSTRUCTION_LINES: tuple[str, ...] = (
    "【会员批量导入模板 — 填写说明】",
    "",
    "1. 请在「会员数据」工作表填写，表头行请勿修改。",
    "2. 带 * 的列为必填：姓名、手机号、套餐类型、配送地址、剩余次数。",
    "3. 套餐类型仅支持：周卡、月卡（与会员档案库口径一致）。",
    "4. 剩余次数：当前剩余可配送餐次；套餐总次数不填时，周卡默认 6、月卡默认 24。",
    "5. 每日份数：每个配送日送达份数，不填默认为 1。",
    "6. 起送日期：格式 YYYY-MM-DD；不填表示无起送限制。",
    "7. 是否自提：填「是」或「否」；自提会员地址可填「自提」。",
    "8. 是否暂停配送：填「是」表示暂不进入配送大表（如「先不送」名单）。",
    "9. 同一手机号在系统中已存在时将自动跳过，不会覆盖已有档案。",
    "10. 填写完成后，在管理后台「会员档案」页点击「导入会员」上传本文件。",
)


def build_member_import_template_xlsx() -> bytes:
    """生成带说明与示例行的导入模板 xlsx 字节流。"""
    wb = Workbook()

    # —— 说明页 ——
    ws_help = wb.active
    ws_help.title = "填写说明"
    title_font = Font(bold=True, size=12)
    for i, line in enumerate(_INSTRUCTION_LINES, start=1):
        cell = ws_help.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = title_font
    ws_help.column_dimensions["A"].width = 88

    # —— 数据页 ——
    ws_data = wb.create_sheet("会员数据")
    header_fill = PatternFill("solid", fgColor="E8F4FD")
    header_font = Font(bold=True)
    for col_idx, title in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_data.column_dimensions[get_column_letter(col_idx)].width = max(12, len(title) * 2 + 2)

    # 示例行（第 2 行，用户导入前可自行删除）
    for col_idx, val in enumerate(_SAMPLE_ROW, start=1):
        ws_data.cell(row=2, column=col_idx, value=val)

    # 打开文件时默认展示数据页
    wb.active = ws_data

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
