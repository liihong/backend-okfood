"""会员批量导入：Excel 模板列解析与行级校验（与 scripts/csv_members_to_sql.py 字段语义对齐）。"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.models.enums import PlanType

# 模板表头（第一行必须包含这些列名，顺序不限）
TEMPLATE_HEADERS: tuple[str, ...] = (
    "姓名",
    "手机号",
    "套餐类型",
    "配送地址",
    "剩余次数",
    "套餐总次数",
    "每日份数",
    "起送日期",
    "是否自提",
    "是否暂停配送",
    "备注",
)

_REQUIRED_HEADERS: frozenset[str] = frozenset(
    {"姓名", "手机号", "套餐类型", "配送地址", "剩余次数"}
)


@dataclass(frozen=True)
class RawMemberImportRow:
    """从 Excel 读取的单行原始数据。"""

    row_no: int
    name: str
    phone_raw: str
    plan_type_raw: str
    address: str
    balance_raw: str
    meal_quota_total_raw: str
    daily_meal_units_raw: str
    delivery_start_raw: str
    store_pickup_raw: str
    delivery_deferred_raw: str
    remarks: str


def normalize_phone(raw: str) -> str | None:
    """规范化手机号为 11 位；无效返回 None。"""
    if raw is None:
        return None
    t = str(raw).strip().replace(" ", "").replace("-", "")
    if not t or t in ("无", "暂无", "—", "-"):
        return None
    if re.fullmatch(r"1\d{10}", t):
        return t
    if re.fullmatch(r"\d{10}", t):
        return "1" + t if not t.startswith("1") else t
    # Excel 可能将手机号读成浮点
    if re.fullmatch(r"\d+\.0", t):
        t = t.split(".")[0]
        if re.fullmatch(r"1\d{10}", t):
            return t
    return None


def parse_int_cell(raw: Any) -> int | None:
    """解析整数单元格；空值返回 None。"""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float):
        if raw != raw:  # NaN
            return None
        return int(raw)
    t = str(raw).strip()
    if not t:
        return None
    try:
        return int(float(t))
    except ValueError:
        return None


def parse_yes_no(raw: Any, *, default: bool = False) -> bool:
    """解析「是/否」「Y/N」「1/0」等布尔语义。"""
    if raw is None:
        return default
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, (int, float)):
        return int(raw) != 0
    t = str(raw).strip().lower()
    if not t:
        return default
    if t in ("是", "yes", "y", "true", "1", "对", "√"):
        return True
    if t in ("否", "no", "n", "false", "0", "×"):
        return False
    return default


def parse_start_date(raw: Any) -> date | None:
    """解析起送日期；支持 YYYY-MM-DD、YYYY/MM/DD 与 Excel 日期序列。"""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    t = str(raw).strip()
    if not t:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(t, fmt).date()
        except ValueError:
            continue
    try:
        n = float(t)
        if 30000 <= n <= 60000:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(n))).date()
    except ValueError:
        pass
    return None


def parse_plan_type(raw: str) -> PlanType | None:
    """解析套餐类型；仅支持周卡/月卡（档案库口径）。"""
    if not raw:
        return None
    t = str(raw).strip()
    if t in (PlanType.WEEK.value, PlanType.MONTH.value):
        return PlanType(t)
    return None


def default_meal_quota_total(plan_type: PlanType, balance: int, explicit: int | None) -> int:
    """套餐总次数：优先显式填写，否则按卡型默认，再否则取剩余次数。"""
    if explicit is not None and explicit >= 0:
        return explicit
    if plan_type == PlanType.MONTH:
        return 24
    if plan_type == PlanType.WEEK:
        return 6
    return max(balance, 0)


def is_store_pickup_address(address: str, store_pickup_flag: bool) -> bool:
    """判断是否为自提：显式标记或地址为自提类文案。"""
    if store_pickup_flag:
        return True
    a = (address or "").strip()
    if a in ("自提", "店内自取", "门店自取"):
        return True
    if "店内自取" in a and len(a) <= 20:
        return True
    return False


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _header_index_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    """表头列名 → 0-based 列索引。"""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        name = _cell_str(cell)
        if name:
            mapping[name] = idx
    return mapping


def parse_member_import_xlsx(file_bytes: bytes) -> tuple[list[RawMemberImportRow], list[str]]:
    """
    解析上传的 xlsx 文件。

    返回 (数据行列表, 文件级错误列表)。文件级错误非空时数据行为空。
    """
    errors: list[str] = []
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return [], ["无法读取 Excel 文件，请使用 .xlsx 格式并确认文件未损坏"]

    try:
        ws = wb.active
        if ws is None:
            return [], ["Excel 工作表为空"]
        # 优先读取「会员数据」工作表（模板含说明页时避免误读）
        if "会员数据" in wb.sheetnames:
            ws = wb["会员数据"]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], ["Excel 无表头行"]

        col_map = _header_index_map(header_row)
        missing = _REQUIRED_HEADERS - set(col_map.keys())
        if missing:
            return [], [f"缺少必填列：{', '.join(sorted(missing))}"]

        def col(row: tuple[Any, ...], name: str) -> str:
            idx = col_map.get(name)
            if idx is None or idx >= len(row):
                return ""
            return _cell_str(row[idx])

        result: list[RawMemberImportRow] = []
        for excel_row_idx, row in enumerate(rows_iter, start=2):
            # 跳过整行空白
            if not any(_cell_str(c) for c in row):
                continue
            result.append(
                RawMemberImportRow(
                    row_no=excel_row_idx,
                    name=col(row, "姓名"),
                    phone_raw=col(row, "手机号"),
                    plan_type_raw=col(row, "套餐类型"),
                    address=col(row, "配送地址"),
                    balance_raw=col(row, "剩余次数"),
                    meal_quota_total_raw=col(row, "套餐总次数"),
                    daily_meal_units_raw=col(row, "每日份数"),
                    delivery_start_raw=col(row, "起送日期"),
                    store_pickup_raw=col(row, "是否自提"),
                    delivery_deferred_raw=col(row, "是否暂停配送"),
                    remarks=col(row, "备注"),
                )
            )
        return result, errors
    finally:
        wb.close()


def normalize_raw_row(raw: RawMemberImportRow) -> tuple[dict[str, Any] | None, list[str]]:
    """
    将原始行转为标准化字段 dict；失败时返回 (None, 错误列表)。
    dict 键与 MemberImportRowData 字段一致。
    """
    msgs: list[str] = []

    name = re.sub(r"\s+", " ", (raw.name or "").strip())
    if not name:
        msgs.append("姓名不能为空")

    phone = normalize_phone(raw.phone_raw)
    if not phone:
        msgs.append("手机号无效，须为 11 位中国大陆手机号")

    plan_type = parse_plan_type(raw.plan_type_raw)
    if plan_type is None:
        msgs.append("套餐类型须填写「周卡」或「月卡」")

    address = (raw.address or "").strip()
    store_pickup_flag = parse_yes_no(raw.store_pickup_raw, default=False)
    store_pickup = is_store_pickup_address(address, store_pickup_flag)
    if not address and not store_pickup:
        msgs.append("配送地址不能为空（自提请填「自提」或「是否自提」填是）")

    balance = parse_int_cell(raw.balance_raw)
    if balance is None:
        msgs.append("剩余次数须为整数")
    elif balance < 0:
        msgs.append("剩余次数不能为负数")

    meal_quota_total = parse_int_cell(raw.meal_quota_total_raw)
    if meal_quota_total is not None and meal_quota_total < 0:
        msgs.append("套餐总次数不能为负数")

    daily_meal_units = parse_int_cell(raw.daily_meal_units_raw)
    if daily_meal_units is None:
        daily_meal_units = 1
    elif not (1 <= daily_meal_units <= 50):
        msgs.append("每日份数须在 1～50 之间")

    delivery_start = parse_start_date(raw.delivery_start_raw)
    if raw.delivery_start_raw.strip() and delivery_start is None:
        msgs.append("起送日期格式无效，请使用 YYYY-MM-DD")

    delivery_deferred = parse_yes_no(raw.delivery_deferred_raw, default=False)
    remarks = (raw.remarks or "").strip() or None
    if remarks and len(remarks) > 500:
        remarks = remarks[:500]

    if msgs:
        return None, msgs

    assert phone is not None and plan_type is not None and balance is not None
    quota = default_meal_quota_total(plan_type, balance, meal_quota_total)

    return (
        {
            "phone": phone,
            "name": name,
            "plan_type": plan_type,
            "address": address if address else "自提",
            "balance": balance,
            "meal_quota_total": quota,
            "daily_meal_units": daily_meal_units,
            "delivery_start_date": delivery_start,
            "store_pickup": store_pickup,
            "delivery_deferred": delivery_deferred,
            "remarks": remarks,
        },
        [],
    )
