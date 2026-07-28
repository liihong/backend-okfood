"""会员批量导入：解析、预览与入库单元测试。"""

from __future__ import annotations

from datetime import time
from io import BytesIO

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.db.base import Base
from app.models.enums import PlanType
from app.models.member import Member
from app.models.member_address import MemberAddress
from app.models.store import Store
from app.models.tenant import Tenant
from app.schemas.member_import import MemberImportConfirmIn
from app.services.admin.member_import_parser import TEMPLATE_HEADERS, normalize_phone, parse_member_import_xlsx
from app.services.admin.member_import_service import build_member_import_preview, confirm_member_import
from app.services.admin.member_import_xlsx import build_member_import_template_xlsx


def _build_xlsx_rows(rows: list[tuple]) -> bytes:
    """辅助：按模板表头构造 xlsx。"""
    wb = Workbook()
    ws = wb.active
    for col, h in enumerate(TEMPLATE_HEADERS, start=1):
        ws.cell(row=1, column=col, value=h)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def import_db() -> Session:
    engine = create_engine("sqlite:///:memory:")
    tables = [Tenant.__table__, Store.__table__, Member.__table__, MemberAddress.__table__]
    Base.metadata.create_all(engine, tables=tables)
    SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)
    session = SessionLocal()
    try:
        session.add_all(
            [
                Tenant(id=3, name="租户3", is_active=True),
                Store(id=1, tenant_id=3, name="门店1", leave_deadline_time=time(21, 0), is_active=True),
            ]
        )
        session.commit()
        yield session
    finally:
        session.close()
        engine.dispose()


def test_normalize_phone():
    assert normalize_phone("13800138000") == "13800138000"
    assert normalize_phone("13800138000.0") == "13800138000"
    assert normalize_phone("无效") is None


def test_template_xlsx_has_headers():
    raw = build_member_import_template_xlsx()
    rows, errs = parse_member_import_xlsx(raw)
    assert not errs
    assert len(rows) >= 1


def test_preview_ready_and_skip(import_db: Session):
    import_db.add(
        Member(
            tenant_id=3,
            store_id=1,
            phone="13900000001",
            name="已存在",
            balance=1,
            meal_quota_total=6,
            plan_type=PlanType.WEEK.value,
            is_active=True,
        )
    )
    import_db.commit()

    xlsx = _build_xlsx_rows(
        [
            ("张三", "13800138001", "周卡", "新乡市测试路1号", 4, 6, 1, "2026-07-01", "否", "否", ""),
            ("李四", "13900000001", "月卡", "新乡市测试路2号", 10, 24, 1, "", "否", "否", ""),
        ]
    )
    preview = build_member_import_preview(import_db, file_bytes=xlsx, tenant_id=3, store_id=1)
    assert preview.summary.total == 2
    assert preview.summary.ready == 1
    assert preview.summary.skip == 1
    assert preview.summary.error == 0


def test_confirm_insert_member(import_db: Session, monkeypatch):
    # 单测环境无 delivery_regions 表，跳过地理编码与划区
    monkeypatch.setattr("app.services.admin.member_import_service.amap.geocode_address", lambda _addr: None)

    xlsx = _build_xlsx_rows(
        [("王五", "13800138002", "周卡", "新乡市测试路3号", 5, 6, 2, "2026-07-15", "否", "否", "备注")]
    )
    preview = build_member_import_preview(import_db, file_bytes=xlsx, tenant_id=3, store_id=1)
    ready_rows = [r.data for r in preview.rows if r.status == "ready" and r.data]
    assert len(ready_rows) == 1

    result = confirm_member_import(
        import_db,
        body=MemberImportConfirmIn(rows=ready_rows),
        tenant_id=3,
        store_id=1,
        operator="admin_test",
    )
    assert result.inserted == 1
    m = import_db.query(Member).filter(Member.phone == "13800138002").one()
    assert m.name == "王五"
    assert m.balance == 5
    assert m.daily_meal_units == 2
    assert m.tenant_id == 3
    addr = import_db.query(MemberAddress).filter(MemberAddress.member_id == m.id).one()
    assert "测试路3号" in (addr.map_location_text or "")


def test_preview_invalid_phone(import_db: Session):
    xlsx = _build_xlsx_rows([("赵六", "123", "周卡", "地址", 1, 6, 1, "", "否", "否", "")])
    preview = build_member_import_preview(import_db, file_bytes=xlsx, tenant_id=3, store_id=1)
    assert preview.summary.error == 1
    assert preview.rows[0].status == "error"
