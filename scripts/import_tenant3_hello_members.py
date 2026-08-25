"""
租户 3（许昌禹州-Hello轻厨）历史会员一次性导入。

数据源：运营台账 xlsx（默认桌面「6月1.xlsx」）。
规则与卡包模版对齐：
- 月午餐/周午餐/午餐 → 仅午餐次数池
- 晚餐 → 仅晚餐次数池
- 午餐+晚餐 → 同一剩余次数同时写入午餐池和晚餐池（模版 meals_grant 入账口径）
- 8月余为当前剩余；忌口等进备注；无手机号跳过
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

_BACKEND_ROOT = Path(__file__).resolve().parent.parent
if str(_BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(_BACKEND_ROOT))

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.constants import STUB_MEMBER_NAME
from app.core.timeutil import beijing_now_naive
from app.db.session import SessionLocal
from app.models.balance_log import BalanceLog
from app.models.enums import BalanceReason, MealPeriod, PlanType
from app.models.member import Member
from app.models.member_meal_period_state import MemberMealPeriodState
from app.services.admin.member_import_parser import normalize_phone, parse_int_cell
from app.services.meal_period.balance import ensure_dinner_period_state, sync_member_is_active_from_period_balances
from app.services.member.member_address_service import upsert_default_address_after_register
from app.services.shared import amap
from app.services.shared.region_assignment import assign_region_for_coords

TENANT_ID = 3
STORE_ID = 3
OPERATOR = "tenant3_xlsx_import"
GEO_PREFIX = "河南省许昌市禹州市"
STORE_LNG = 113.485717
STORE_LAT = 34.149193
REGION_NAMES = frozenset({"东", "西", "南", "北", "东南", "东北", "西南", "西北"})


@dataclass
class ParsedRow:
    """xlsx 一行解析结果。"""

    row_no: int
    name: str
    wechat: str
    phone: str | None
    address: str
    region: str
    pickup: bool
    units: int
    period: str  # lunch / dinner / both
    plan_type: str
    remaining: int
    remark: str
    skip_reason: str | None = None


@dataclass
class MemberDraft:
    """按手机号合并后的待写入档案。"""

    phone: str
    name: str
    wechat_name: str | None
    address: str
    pickup: bool
    lunch_balance: int = 0
    dinner_balance: int = 0
    lunch_units: int = 1
    dinner_units: int = 1
    plan_type: str = PlanType.MONTH.value
    has_combo: bool = False
    remarks: str | None = None
    row_nos: list[int] = field(default_factory=list)
    name_conflict: str | None = None


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _parse_units(raw: str) -> int:
    t = (raw or "").strip()
    if not t:
        return 1
    m = re.search(r"(\d+)", t)
    if m:
        n = int(m.group(1))
        if 1 <= n <= 50:
            return n
    return 1


def _is_pickup(raw: str, address: str) -> bool:
    t = (raw or "").strip()
    if t in ("是", "自取", "自提", "店内自取", "门店自取"):
        return True
    a = (address or "").strip()
    if a in ("自提", "自取", "店内自取", "门店自取"):
        return True
    return False


def _parse_meal(raw: str) -> tuple[str, str] | None:
    """返回 (period, plan_type)；无法识别则 None。"""
    t = re.sub(r"\s+", "", raw or "")
    if not t:
        return None
    plan = PlanType.WEEK.value if "周" in t else PlanType.MONTH.value
    if "晚" in t and "午" in t:
        return "both", plan
    if "晚" in t:
        return "dinner", plan
    if "午" in t:
        return "lunch", plan
    return None


def _merge_remark(*parts: str) -> str | None:
    out: list[str] = []
    for p in parts:
        s = re.sub(r"\s+", " ", (p or "").replace("\n", " ")).strip()
        if s and s not in out:
            out.append(s)
    return "；".join(out) if out else None


def parse_xlsx(path: Path) -> tuple[list[ParsedRow], list[str]]:
    """读取台账，返回解析行与文件级错误。"""
    wb = load_workbook(filename=path, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows: list[ParsedRow] = []
        empty_streak = 0
        # 该表曾带百万空行格式，连续空行即停
        for r in range(2, min(int(ws.max_row or 2) + 1, 2000)):
            wechat = _cell_str(ws.cell(r, 1).value)
            name = re.sub(r"\s+", " ", _cell_str(ws.cell(r, 2).value))
            phone_raw = _cell_str(ws.cell(r, 3).value)
            address = _cell_str(ws.cell(r, 4).value)
            region = _cell_str(ws.cell(r, 5).value)
            pickup_raw = _cell_str(ws.cell(r, 6).value)
            units_raw = _cell_str(ws.cell(r, 7).value)
            meal_raw = _cell_str(ws.cell(r, 8).value)
            remaining = parse_int_cell(ws.cell(r, 45).value)
            remark = _cell_str(ws.cell(r, 78).value)
            if not (name or wechat or phone_raw or meal_raw):
                empty_streak += 1
                if empty_streak >= 20:
                    break
                continue
            empty_streak = 0
            pickup = _is_pickup(pickup_raw, address)
            parsed_meal = _parse_meal(meal_raw)
            skip: str | None = None
            phone = normalize_phone(phone_raw)
            if not phone:
                skip = "无有效手机号"
            elif parsed_meal is None:
                skip = f"餐别无法识别：{meal_raw or '（空）'}"
            elif remaining is None:
                skip = "缺少 8月余"
            elif remaining < 0:
                remaining = 0
            if skip is None and not address and not pickup:
                skip = "无配送地址且非自提"
            period, plan_type = parsed_meal if parsed_meal else ("lunch", PlanType.MONTH.value)
            rows.append(
                ParsedRow(
                    row_no=r,
                    name=name or wechat or "未命名",
                    wechat=wechat,
                    phone=phone,
                    address=address if address else ("自提" if pickup else ""),
                    region=region if region in REGION_NAMES else "",
                    pickup=pickup,
                    units=_parse_units(units_raw),
                    period=period,
                    plan_type=plan_type,
                    remaining=int(remaining or 0),
                    remark=_merge_remark(
                        remark,
                        f"片区{region}" if region and region in REGION_NAMES else "",
                        region if region and region not in REGION_NAMES else "",
                    )
                    or "",
                    skip_reason=skip,
                )
            )
        return rows, []
    finally:
        wb.close()


def merge_by_phone(rows: list[ParsedRow]) -> tuple[list[MemberDraft], list[dict[str, Any]]]:
    """同一手机号合并；不同姓名时保留剩余次数合计更高的姓名。"""
    skipped: list[dict[str, Any]] = []
    groups: dict[str, list[ParsedRow]] = {}
    order: list[str] = []
    for row in rows:
        if row.skip_reason:
            skipped.append(
                {"row": row.row_no, "name": row.name, "phone": row.phone, "reason": row.skip_reason}
            )
            continue
        assert row.phone is not None
        if row.phone not in groups:
            order.append(row.phone)
            groups[row.phone] = []
        groups[row.phone].append(row)

    drafts: list[MemberDraft] = []
    for phone in order:
        items = groups[phone]
        lunch = dinner = 0
        lunch_units = dinner_units = 1
        has_combo = False
        pickup = False
        plan = PlanType.MONTH.value
        remarks: list[str] = []
        address = ""
        wechat = ""
        # 姓名：按该行贡献的次数加权，避免路遥/马佳星这类撞号用错名
        name_score: dict[str, int] = {}
        for it in items:
            contrib = it.remaining
            if it.period == "both":
                contrib *= 2
            name_score[it.name] = name_score.get(it.name, 0) + contrib
            if it.wechat and not wechat:
                wechat = it.wechat
            if it.address and (not address or it.pickup is False):
                address = it.address
            pickup = pickup or it.pickup
            if it.plan_type == PlanType.WEEK.value and plan != PlanType.MONTH.value:
                plan = PlanType.WEEK.value
            if it.plan_type == PlanType.MONTH.value:
                plan = PlanType.MONTH.value
            if it.period == "lunch":
                lunch += it.remaining
                lunch_units = max(lunch_units, it.units)
            elif it.period == "dinner":
                dinner += it.remaining
                dinner_units = max(dinner_units, it.units)
            else:
                has_combo = True
                lunch += it.remaining
                dinner += it.remaining
                lunch_units = max(lunch_units, it.units)
                dinner_units = max(dinner_units, it.units)
            if it.remark:
                remarks.append(it.remark)
        name = max(name_score.items(), key=lambda kv: kv[1])[0]
        conflict = None
        names = {it.name for it in items}
        if len(names) > 1:
            conflict = "同号不同名：" + " / ".join(sorted(names))
            remarks.append(conflict)
        if pickup and not address:
            address = "自提"
        drafts.append(
            MemberDraft(
                phone=phone,
                name=name[:100],
                wechat_name=(wechat or name)[:100] if (wechat or name) else None,
                address=address[:500] if address else "自提",
                pickup=pickup,
                lunch_balance=max(0, lunch),
                dinner_balance=max(0, dinner),
                lunch_units=lunch_units,
                dinner_units=dinner_units,
                plan_type=plan,
                has_combo=has_combo,
                remarks=_merge_remark(*remarks),
                row_nos=[it.row_no for it in items],
                name_conflict=conflict,
            )
        )
    return drafts, skipped


def _quota(balance: int, default_total: int) -> int:
    return max(int(balance), int(default_total), 0)


def _geocode_line(address: str, pickup: bool) -> tuple[float | None, float | None]:
    if pickup:
        return STORE_LNG, STORE_LAT
    raw = (address or "").strip()
    if not raw:
        return None, None
    if any(k in raw for k in ("河南", "许昌", "禹州")):
        query = raw
    else:
        query = f"{GEO_PREFIX}{raw}"
    coords = amap.geocode_address(query)
    if coords:
        return coords[0], coords[1]
    return None, None


def _existing_map(db: Session) -> dict[str, Member]:
    rows = db.scalars(
        select(Member).where(
            Member.store_id == STORE_ID,
            Member.deleted_at.is_(None),
        )
    ).all()
    return {str(m.phone): m for m in rows}


def _dinner_balance(db: Session, member_id: int) -> int:
    row = db.get(
        MemberMealPeriodState,
        {"member_id": int(member_id), "meal_period": MealPeriod.DINNER.value},
    )
    return int(row.balance or 0) if row is not None else 0


def apply_drafts(
    db: Session,
    drafts: list[MemberDraft],
    *,
    geocode: bool,
) -> dict[str, Any]:
    """写入会员、地址、晚餐次数池与次数流水。"""
    existing = _existing_map(db)
    inserted = updated = skipped = 0
    messages: list[str] = []
    geo_ok = geo_fail = 0
    geo_cache: dict[str, tuple[float | None, float | None]] = {}

    for draft in drafts:
        old = existing.get(draft.phone)
        if old is not None:
            old_dinner = _dinner_balance(db, int(old.id))
            old_lunch = int(old.balance or 0)
            is_stub = (old.name or "") == STUB_MEMBER_NAME or (old_lunch == 0 and old_dinner == 0)
            if not is_stub and (old_lunch > 0 or old_dinner > 0):
                skipped += 1
                messages.append(f"跳过已有次数会员 {draft.phone} {old.name}（午餐{old_lunch}/晚餐{old_dinner}）")
                continue

        lng = lat = region_id = None
        if not draft.pickup:
            key = draft.address
            if geocode:
                if key not in geo_cache:
                    geo_cache[key] = _geocode_line(draft.address, False)
                    time.sleep(0.12)
                lng, lat = geo_cache[key]
            if lng is not None and lat is not None:
                geo_ok += 1
                r = assign_region_for_coords(db, lng, lat, tenant_id=TENANT_ID)
                region_id = int(r.id) if r else None
            else:
                geo_fail += 1
        else:
            lng, lat = STORE_LNG, STORE_LAT

        lunch_default = 48 if draft.has_combo else (6 if draft.plan_type == PlanType.WEEK.value else 24)
        dinner_default = 48 if draft.has_combo else 24
        lunch_quota = _quota(draft.lunch_balance, lunch_default) if draft.lunch_balance else 0
        dinner_quota = _quota(draft.dinner_balance, dinner_default) if draft.dinner_balance else 0

        if old is None:
            member = Member(
                tenant_id=TENANT_ID,
                store_id=STORE_ID,
                phone=draft.phone,
                name=draft.name,
                wechat_name=draft.wechat_name,
                remarks=draft.remarks,
                balance=draft.lunch_balance,
                daily_meal_units=draft.lunch_units,
                meal_quota_total=lunch_quota,
                plan_type=draft.plan_type,
                is_active=False,
                delivery_start_date=None,
                delivery_deferred=False,
                store_pickup=draft.pickup,
            )
            db.add(member)
            db.flush()
            inserted += 1
            action = "新建"
        else:
            member = old
            member.name = draft.name
            member.wechat_name = draft.wechat_name
            member.remarks = draft.remarks
            member.balance = draft.lunch_balance
            member.daily_meal_units = draft.lunch_units
            member.meal_quota_total = lunch_quota
            member.plan_type = draft.plan_type
            member.store_pickup = draft.pickup
            member.delivery_deferred = False
            member.deleted_at = None
            member.updated_at = beijing_now_naive()
            updated += 1
            action = "更新占位档"

        if draft.dinner_balance > 0 or draft.has_combo:
            drow = ensure_dinner_period_state(db, int(member.id))
            drow.balance = draft.dinner_balance
            drow.daily_meal_units = draft.dinner_units
            drow.meal_quota_total = dinner_quota
            db.add(drow)
        sync_member_is_active_from_period_balances(db, member)

        upsert_default_address_after_register(
            db,
            member_id=int(member.id),
            contact_name=draft.name,
            contact_phone=draft.phone,
            address_line=draft.address,
            remarks=draft.remarks,
            delivery_region_id=region_id,
            lng=lng,
            lat=lat,
        )

        if draft.lunch_balance:
            db.add(
                BalanceLog(
                    member_id=int(member.id),
                    meal_period=MealPeriod.LUNCH.value,
                    change=int(draft.lunch_balance),
                    reason=BalanceReason.ADMIN_ADJUST.value,
                    operator=OPERATOR,
                    detail=f"{action}导入午餐剩余{draft.lunch_balance}次；xlsx行{','.join(str(x) for x in draft.row_nos)}",
                )
            )
        if draft.dinner_balance:
            db.add(
                BalanceLog(
                    member_id=int(member.id),
                    meal_period=MealPeriod.DINNER.value,
                    change=int(draft.dinner_balance),
                    reason=BalanceReason.ADMIN_ADJUST.value,
                    operator=OPERATOR,
                    detail=f"{action}导入晚餐剩余{draft.dinner_balance}次；xlsx行{','.join(str(x) for x in draft.row_nos)}",
                )
            )

    db.commit()
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped_existing": skipped,
        "geo_ok": geo_ok,
        "geo_fail": geo_fail,
        "messages": messages,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="导入租户3 Hello轻厨历史会员")
    parser.add_argument(
        "--file",
        default=r"C:\Users\Administrator\Desktop\6月1.xlsx",
        help="xlsx 路径",
    )
    parser.add_argument("--apply", action="store_true", help="真正写入数据库；默认只预览")
    parser.add_argument("--no-geocode", action="store_true", help="跳过高德地理编码")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.is_file():
        print(f"文件不存在: {path}")
        return 1

    rows, file_errors = parse_xlsx(path)
    if file_errors:
        print("文件错误:", file_errors)
        return 1
    drafts, skipped = merge_by_phone(rows)
    preview = {
        "source_rows": len(rows),
        "drafts": len(drafts),
        "lunch_only": sum(1 for d in drafts if d.lunch_balance > 0 and d.dinner_balance == 0),
        "dinner_only": sum(1 for d in drafts if d.dinner_balance > 0 and d.lunch_balance == 0),
        "both": sum(1 for d in drafts if d.lunch_balance > 0 and d.dinner_balance > 0),
        "pickup": sum(1 for d in drafts if d.pickup),
        "name_conflicts": [
            {"phone": d.phone, "keep_name": d.name, "detail": d.name_conflict}
            for d in drafts
            if d.name_conflict
        ],
        "skipped_rows": skipped,
        "lunch_sum": sum(d.lunch_balance for d in drafts),
        "dinner_sum": sum(d.dinner_balance for d in drafts),
        "samples": [
            {
                "phone": d.phone,
                "name": d.name,
                "lunch": d.lunch_balance,
                "dinner": d.dinner_balance,
                "units": d.lunch_units,
                "pickup": d.pickup,
                "plan": d.plan_type,
                "addr": d.address,
            }
            for d in drafts[:8]
        ],
    }
    out_path = _BACKEND_ROOT / "_tenant3_import_preview.json"
    out_path.write_text(json.dumps(preview, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"解析完成：源行 {preview['source_rows']}，合并后 {preview['drafts']} 人，"
        f"仅午 {preview['lunch_only']}，仅晚 {preview['dinner_only']}，午+晚 {preview['both']}，"
        f"跳过源行 {len(skipped)}"
    )
    print(f"预览已写 {out_path}")
    if not args.apply:
        print("未加 --apply，不写库。")
        return 0

    db = SessionLocal()
    try:
        result = apply_drafts(db, drafts, geocode=not args.no_geocode)
        result_path = _BACKEND_ROOT / "_tenant3_import_result.json"
        result_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
        print(
            f"入库完成：新建 {result['inserted']}，更新 {result['updated']}，"
            f"跳过已有 {result['skipped_existing']}，"
            f"地理编码成功 {result['geo_ok']} / 失败 {result['geo_fail']}"
        )
        for msg in result["messages"]:
            print(msg)
        return 0
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
