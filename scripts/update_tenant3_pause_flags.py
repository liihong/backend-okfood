"""按台账「暂停」列更新租户 3 会员配送状态。

1=暂停配送，2=请假（无日期则写入今天起一年区间），4=周六不送餐；可按位组合（如 3=1+2）。
"""

from __future__ import annotations

import sys
from collections import defaultdict
from datetime import timedelta
from pathlib import Path

_BACKEND_ROOT = Path(__file__).resolve().parent.parent
if str(_BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(_BACKEND_ROOT))

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.timeutil import now_shanghai, today_shanghai
from app.db.session import SessionLocal
from app.models.enums import LeaveType, MealPeriod
from app.models.member import Member
from app.models.member_meal_period_state import MemberMealPeriodState
from app.services.admin.member_import_parser import normalize_phone
from app.services.meal_period.leave_fields import apply_leave_type_to_row
from app.services.member.member_delivery_state_service import apply_pause_delivery
from app.services.member.member_operation_log_service import (
    OP_ADMIN_UPDATE_SKIP_SATURDAY,
    OP_LEAVE_RANGE,
    OP_PAUSE_DELIVERY,
    record_member_operation,
)

TENANT_ID = 3
STORE_ID = 3
PAUSE_COL = 76
XLSX = Path(r"C:\Users\Administrator\Desktop\6月11111111.xlsx")
OPERATOR = "tenant3_xlsx_pause"
FLAG_PAUSE = 1
FLAG_LEAVE = 2
FLAG_SKIP_SAT = 4


def _parse_flag(raw) -> int:
    if raw is None or str(raw).strip() == "":
        return 0
    try:
        n = int(float(raw))
    except (TypeError, ValueError):
        return 0
    if n < 0:
        return 0
    return n


def load_flags(path: Path) -> dict[str, int]:
    """手机号 → 暂停位标记（同号取按位或）。"""
    wb = load_workbook(filename=path, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        merged: dict[str, int] = defaultdict(int)
        empty_streak = 0
        for r in range(2, 2000):
            name = ws.cell(r, 2).value
            phone_raw = ws.cell(r, 3).value
            pause_raw = ws.cell(r, PAUSE_COL).value
            if name is None and phone_raw is None and pause_raw is None:
                empty_streak += 1
                if empty_streak >= 20:
                    break
                continue
            empty_streak = 0
            phone = normalize_phone(str(phone_raw).strip() if phone_raw is not None else "")
            if not phone:
                continue
            merged[phone] |= _parse_flag(pause_raw)
        return dict(merged)
    finally:
        wb.close()


def main() -> int:
    if not XLSX.is_file():
        print(f"文件不存在: {XLSX}")
        return 1
    flags = load_flags(XLSX)
    today = today_shanghai()
    leave_end = today + timedelta(days=365)
    now = now_shanghai()

    db = SessionLocal()
    stats = {
        "xlsx_phones": len(flags),
        "paused": 0,
        "leave": 0,
        "skip_sat": 0,
        "skipped_zero": 0,
        "missing": 0,
        "missing_phones": [],
    }
    try:
        members = db.scalars(
            select(Member).where(
                Member.tenant_id == TENANT_ID,
                Member.store_id == STORE_ID,
                Member.deleted_at.is_(None),
            )
        ).all()
        by_phone = {str(m.phone): m for m in members}

        for phone, flag in flags.items():
            if flag == 0:
                stats["skipped_zero"] += 1
                continue
            m = by_phone.get(phone)
            if m is None:
                stats["missing"] += 1
                stats["missing_phones"].append(phone)
                continue

            if flag & FLAG_PAUSE:
                before = bool(m.delivery_deferred)
                apply_pause_delivery(db, m)
                if not before:
                    record_member_operation(
                        db,
                        member_id=int(m.id),
                        operation_type=OP_PAUSE_DELIVERY,
                        summary="台账导入：暂停配送",
                        before={"delivery_deferred": before},
                        after={"delivery_deferred": True},
                        operator=OPERATOR,
                        source="admin",
                    )
                stats["paused"] += 1

            if flag & FLAG_LEAVE:
                apply_leave_type_to_row(
                    m,
                    LeaveType.RANGE,
                    today,
                    leave_end,
                    skip_leave_deadline=True,
                    now=now,
                )
                drow = db.get(
                    MemberMealPeriodState,
                    {"member_id": int(m.id), "meal_period": MealPeriod.DINNER.value},
                )
                if drow is not None:
                    apply_leave_type_to_row(
                        drow,
                        LeaveType.RANGE,
                        today,
                        leave_end,
                        skip_leave_deadline=True,
                        now=now,
                    )
                    db.add(drow)
                record_member_operation(
                    db,
                    member_id=int(m.id),
                    operation_type=OP_LEAVE_RANGE,
                    summary=f"台账导入：请假 {today.isoformat()} ~ {leave_end.isoformat()}",
                    after={"leave_range_start": today.isoformat(), "leave_range_end": leave_end.isoformat()},
                    operator=OPERATOR,
                    source="admin",
                )
                stats["leave"] += 1

            if flag & FLAG_SKIP_SAT:
                before = bool(m.skip_subscription_saturday)
                m.skip_subscription_saturday = True
                if not before:
                    record_member_operation(
                        db,
                        member_id=int(m.id),
                        operation_type=OP_ADMIN_UPDATE_SKIP_SATURDAY,
                        summary="台账导入：周六不送餐",
                        before={"skip_subscription_saturday": before},
                        after={"skip_subscription_saturday": True},
                        operator=OPERATOR,
                        source="admin",
                    )
                stats["skip_sat"] += 1

        db.commit()
        print(
            f"xlsx手机号 {stats['xlsx_phones']}，"
            f"暂停 {stats['paused']}，"
            f"请假 {stats['leave']}（{today}~{leave_end}），"
            f"周六不送 {stats['skip_sat']}，"
            f"标记为空未改 {stats['skipped_zero']}，"
            f"库中无此号 {stats['missing']}"
        )
        if stats["missing_phones"]:
            print("未匹配:", ", ".join(stats["missing_phones"][:20]))
        return 0
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
